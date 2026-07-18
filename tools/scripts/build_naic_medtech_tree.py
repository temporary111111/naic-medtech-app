from __future__ import annotations

import json
import re
import unicodedata
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "data" / "source" / "NAIC MEDTECH SYSTEM DATA.xlsx"
OUT_DIR = ROOT / "artifacts" / "schema"
JSON_OUTPUT_PATH = OUT_DIR / "naic_medtech_structure.json"
APP_SCHEMA_OUTPUT_PATH = OUT_DIR / "naic_medtech_app_schema.json"
MARKDOWN_OUTPUT_PATH = OUT_DIR / "naic_medtech_structure.md"
HTML_OUTPUT_PATH = OUT_DIR / "naic_medtech_structure.html"
DIAGRAM_HTML_OUTPUT_PATH = OUT_DIR / "naic_medtech_tree_diagram.html"
APP_SCHEMA_VERSION = "1.0.1"


FORM_SPECS = [
    {
        "sheet": "SEMEN - Clinical Microscopy",
        "group": "Clinical Microscopy",
        "group_kind": "category",
        "form": "Semen",
    },
    {
        "sheet": "URINE- Clinical Microscopy",
        "group": "Clinical Microscopy",
        "group_kind": "category",
        "form": "Urine",
    },
    {
        "sheet": "FECALYSIS - Clinical Microscopy",
        "group": "Clinical Microscopy",
        "group_kind": "category",
        "form": "Fecalysis",
        "merge_repeated_sections": True,
    },
    {
        "sheet": "OGTT - Blood Chemistry",
        "group": "Blood Chemistry",
        "group_kind": "category",
        "form": "OGTT",
        "synthetic_sections": {
            "Additional Tests": {
                "2 HOURS POST PRANDIAL",
                "50 G ORAL GLUCOSE CHALLENGE",
                "Others",
            }
        },
    },
    {
        "sheet": "HBA1C - Blood Chemistry",
        "group": "Blood Chemistry",
        "group_kind": "category",
        "form": "HBA1C",
    },
    {
        "sheet": "BCMALE - Blood Chemistry",
        "group": "Blood Chemistry",
        "group_kind": "category",
        "form": "Male",
        "move_fields_after": [("IONIZED CALCIUM", "CHLORIDE")],
    },
    {
        "sheet": "BCFEMALE - Blood Chemistry",
        "group": "Blood Chemistry",
        "group_kind": "category",
        "form": "Female",
        "move_fields_after": [("IONIZED CALCIUM", "CHLORIDE")],
    },
    {
        "sheet": "SEROLOGY",
        "group": "Serology",
        "group_kind": "category",
        "form": "Serology",
        "synthetic_sections": {
            "Other Serology Tests": {
                "HbsAg SCREENING:",
                "VDRL:",
                "ANTI-HCV:",
                "ASO TITER:",
                "OTHERS",
            }
        },
    },
    {
        "sheet": "CARDIACI - Serology",
        "group": "Serology",
        "group_kind": "category",
        "form": "Cardiaci",
    },
    {
        "sheet": "HIV 1&2 TESTING - Serology",
        "group": "Serology",
        "group_kind": "category",
        "form": "HIV 1&2 Testing",
    },
    {
        "sheet": "COVID 19 ANTIGEN (RAPID TEST) -",
        "group": "Serology",
        "group_kind": "category",
        "form": "COVID 19 Antigen (Rapid Test)",
    },
    {
        "sheet": "BBANK - Blood Bank",
        "group": "Blood Bank",
        "group_kind": "standalone_form",
        "form": "Blood Bank",
    },
    {
        "sheet": "ABG - Blood Gas Analysis",
        "group": "Blood Gas Analysis",
        "group_kind": "standalone_form",
        "form": "Blood Gas Analysis",
    },
    {
        "sheet": "HEMATOLOGY",
        "group": "Hematology",
        "group_kind": "category",
        "form": "Hematology",
    },
    {
        "sheet": "PROTIME, APTT - Hematology",
        "group": "Hematology",
        "group_kind": "category",
        "form": "Pro-Time, APTT",
    },
    {
        "sheet": "MICROBIOLOGY",
        "group": "Microbiology",
        "group_kind": "standalone_form",
        "form": "Microbiology",
    },
]


COMMON_FIELDS = {
    "field",
    "patient information",
    "name",
    "age",
    "sex",
    "date",
    "date/time",
    "requesting physician",
    "room",
    "case number",
    "medical technologist",
    "pathologist",
}


HEADER_ALIASES = {
    "field": "field",
    "inputtype": "input_type",
    "dropdownlistoptions": "options",
    "normalvalue": "normal_value",
    "normalvalues": "normal_value",
    "notes": "notes",
}


CHAR_REPLACEMENTS = str.maketrans(
    {
        "’": "'",
        "“": '"',
        "”": '"',
        "–": "-",
        "—": "-",
        "°": " deg ",
    }
)


REQUESTING_PHYSICIAN_OPTIONS = [
    "DR. RAUL VILLAR",
    "DR. LAVINIA BELTIJAR",
    "DR. MELWANI GARRIDO",
    "DR. MARIANIDA SISANTE",
    "DR. LEONA CARMEN SISANTE",
    "DR. ALMA LOWENA ANACAY",
    "DR. MARIBENZ ANGON",
    "DR. NELSON PIPIT",
    "DR. ELENITA SISAYAN",
    "DR. GIELZEN JOI SISAYAN",
    "DR. PERLITA CASTRO",
    "DR. BAYANI PASCO",
    "DR. DEXTER SCHROTH",
    "DR. JAYMEE SCHROTH",
    "DR. AIKO SALORSANO",
    "DR. LUCILA OBILLO",
    "DR. NOEL OBILLO",
    "DR. ARNEL MILAY",
    "DR. CANARIE JOY ESGUERRA",
    "DR. ELIZABETH PANGANIBAN",
    "DR. DONNALIZA CRUZ",
    "DR. CHERRIE ANN ANGON",
    "DR. HANNA TRISSIA SUMABONG",
    "DR. JERICA CRISTEL ESGUERRA",
    "DR. KENNETH JAVIER",
    "DR. VERONICA ALERTA",
    "DR. ANGELA FERNANDO",
    "DR. LYSSEL SARACANLAO",
    "DR. IDGEE GABRIEL BONDOC",
]


ROOM_OPTIONS = [
    "ST. THOMAS",
    "ST. ANDREW",
    "ST. JOSEPH",
    "ST. FRANCIS",
    "ST. TIMOTHY",
    "ST. PAUL",
    "ST. JOHN",
    "ST. GABRIEL",
    "ST. ANTHONY",
    "ST. PETER",
    "ST. MATTHEW",
    "ST. DOMINIC",
    "ST. AUGUSTINE",
    "ST. JAMES",
    "ST. MICHAEL",
    "ST. LUKE",
    "ST. LOUIE",
    "ST. JUDE",
    "NICU",
    "OPD",
    "ER",
]


def common_option_key(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", ascii_value.lower()).strip("_") or "option"


def build_common_options(field_id: str, values: list[str]) -> list[dict[str, object]]:
    options: list[dict[str, object]] = []
    used: set[str] = set()
    for order, value in enumerate(values, start=1):
        base_key = common_option_key(value)
        key = base_key
        suffix = 2
        while key in used:
            key = f"{base_key}_{suffix}"
            suffix += 1
        used.add(key)
        options.append(
            {
                "id": f"{field_id}.{key}",
                "key": key,
                "name": value,
                "order": order,
            }
        )
    return options


COMMON_FIELD_SET = {
    "id": "default_lab_request",
    "name": "Default Lab Request Metadata",
    "notes": [
        "These shared fields were excluded from the visual tree and moved into one reusable app field set.",
        "The workbook varies between Date and Date/Time, so that field is represented conservatively as date_or_datetime.",
    ],
    "fields": [
        {
            "id": "default_lab_request.name",
            "key": "name",
            "name": "Name",
            "order": 1,
            "control": "input",
            "data_type": "text",
        },
        {
            "id": "default_lab_request.age",
            "key": "age",
            "name": "Age",
            "order": 2,
            "control": "input",
            "data_type": "text",
        },
        {
            "id": "default_lab_request.sex",
            "key": "sex",
            "name": "Sex",
            "order": 3,
            "control": "select",
            "data_type": "enum",
            "options": build_common_options("default_lab_request.sex", ["Male", "Female"]),
        },
        {
            "id": "default_lab_request.date_or_datetime",
            "key": "date_or_datetime",
            "name": "Date / Date-Time",
            "order": 4,
            "control": "input",
            "data_type": "date_or_datetime",
        },
        {
            "id": "default_lab_request.requesting_physician",
            "key": "requesting_physician",
            "name": "Requesting Physician",
            "order": 5,
            "control": "select",
            "data_type": "enum",
            "options": build_common_options(
                "default_lab_request.requesting_physician",
                REQUESTING_PHYSICIAN_OPTIONS,
            ),
        },
        {
            "id": "default_lab_request.room",
            "key": "room",
            "name": "Room",
            "order": 6,
            "control": "select",
            "data_type": "enum",
            "options": build_common_options("default_lab_request.room", ROOM_OPTIONS),
        },
        {
            "id": "default_lab_request.case_number",
            "key": "case_number",
            "name": "Case Number",
            "order": 7,
            "control": "input",
            "data_type": "text",
        },
    ],
}


DEFAULT_SIGNATORIES = [
    {
        "id": "medical_technologist_1",
        "label": "Analyzed by:",
        "designation": "Medical Technologist (RMT)",
        "input_type": "person_dropdown",
        "required": True,
        "show_on_print": True,
        "show_license": True,
        "signature_line": True,
        "default_option_id": "",
        "options": [
            {"id": "imelda_a_elemia", "name": "Imelda A. Elemia, RMT", "license": "0036643", "order": 1},
            {"id": "crystel_c_tesoro", "name": "Crystel C. Tesoro, RMT", "license": "0103760", "order": 2},
            {"id": "ma_jesusa_b_vite", "name": "Ma. Jesusa B. Vite, RMT", "license": "0118710", "order": 3},
            {"id": "andrea_coleen_a_avellones", "name": "Andrea Coleen A. Avellones, RMT", "license": "0119501", "order": 4},
            {"id": "julie_kyle_a_ronato", "name": "Julie Kyle A. Ronato, RMT", "license": "0119616", "order": 5},
            {"id": "shiela_mae_d_libradilla", "name": "Shiela Mae D. Libradilla, RMT", "license": "0135995", "order": 6},
        ],
    },
    {
        "id": "medical_technologist_2",
        "label": "Verified by:",
        "designation": "Medical Technologist (RMT)",
        "input_type": "person_dropdown",
        "required": True,
        "show_on_print": True,
        "show_license": True,
        "signature_line": True,
        "default_option_id": "",
        "options": [
            {"id": "imelda_a_elemia", "name": "Imelda A. Elemia, RMT", "license": "0036643", "order": 1},
            {"id": "crystel_c_tesoro", "name": "Crystel C. Tesoro, RMT", "license": "0103760", "order": 2},
            {"id": "ma_jesusa_b_vite", "name": "Ma. Jesusa B. Vite, RMT", "license": "0118710", "order": 3},
            {"id": "andrea_coleen_a_avellones", "name": "Andrea Coleen A. Avellones, RMT", "license": "0119501", "order": 4},
            {"id": "julie_kyle_a_ronato", "name": "Julie Kyle A. Ronato, RMT", "license": "0119616", "order": 5},
            {"id": "shiela_mae_d_libradilla", "name": "Shiela Mae D. Libradilla, RMT", "license": "0135995", "order": 6},
        ],
    },
    {
        "id": "pathologist",
        "label": "Noted by:",
        "designation": "Pathologist",
        "input_type": "stamp_image",
        "required": False,
        "show_on_print": True,
        "show_license": False,
        "signature_line": True,
        "default_option_id": "",
        "stamp_image_url": "/signatory-stamps/default-pathologist-stamp.png",
        "stamp_image_filename": "default-pathologist-stamp.png",
        "stamp_image_mime_type": "image/png",
        "options": [],
    },
]


VITAL_SIGNS_CHILD_SPECS = {
    "blood pressure": {
        "data_type": "text",
        "unit_hint": "mmHg",
    },
    "pulse rate": {
        "data_type": "number",
        "unit_hint": "bpm",
    },
    "respiratory rate": {
        "data_type": "number",
        "unit_hint": "cpm",
    },
    "temperature": {
        "data_type": "number",
        "unit_hint": "deg C",
    },
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = text.translate(CHAR_REPLACEMENTS)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return text.strip()


def collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def canonical_key(value: str) -> str:
    normalized = collapse_spaces(clean_text(value)).rstrip(":")
    normalized = normalized.replace("’", "'")
    return normalized.lower()


def display_label(value: str) -> str:
    return collapse_spaces(clean_text(value)).rstrip(":")


def slugify(value: str) -> str:
    normalized = clean_text(value).replace("&", " and ").lower()
    slug = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return slug or "item"


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def is_placeholder(value: str) -> bool:
    compact = collapse_spaces(clean_text(value)).replace(" ", "")
    return bool(compact) and set(compact) == {"-"}


def parse_option_lines(value: str) -> list[str]:
    raw = clean_text(value)
    if not raw or is_placeholder(raw):
        return []

    options = []
    for line in raw.split("\n"):
        item = collapse_spaces(line)
        if item and not is_placeholder(item):
            options.append(item)
    return options


def append_unique(items: list[str], value: str) -> None:
    if value and value not in items:
        items.append(value)


def combine_notes(*values: str) -> list[str]:
    notes: list[str] = []
    for value in values:
        text = collapse_spaces(clean_text(value))
        if text and not is_placeholder(text):
            append_unique(notes, text)
    return notes


def get_header_map(worksheet) -> tuple[dict[str, int], list[str]]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map: dict[str, int] = {}
    extra_notes: list[str] = []

    for index, cell in enumerate(first_row):
        raw = clean_text(cell)
        if not raw:
            continue
        alias = HEADER_ALIASES.get(normalize_header(raw))
        if alias:
            header_map[alias] = index
        else:
            append_unique(extra_notes, collapse_spaces(raw))

    return header_map, extra_notes


def extract_rows(worksheet) -> tuple[list[dict[str, str | int]], list[str]]:
    header_map, extra_notes = get_header_map(worksheet)
    rows: list[dict[str, str | int]] = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        field = clean_text(row[header_map["field"]]) if "field" in header_map else ""
        input_type = clean_text(row[header_map["input_type"]]) if "input_type" in header_map else ""
        options = clean_text(row[header_map["options"]]) if "options" in header_map else ""
        normal_value = clean_text(row[header_map["normal_value"]]) if "normal_value" in header_map else ""
        notes = clean_text(row[header_map["notes"]]) if "notes" in header_map else ""

        if not any((field, input_type, options, normal_value, notes)):
            continue

        rows.append(
            {
                "row": row_index,
                "field": field,
                "input_type": input_type,
                "options": options,
                "normal_value": normal_value,
                "notes": notes,
            }
        )

    return rows, extra_notes


def build_field(row: dict[str, str | int]) -> dict[str, object]:
    source_label = clean_text(row["field"])
    field = {
        "name": display_label(source_label),
        "sheet_row": row["row"],
    }

    if field["name"] != source_label:
        field["source_label"] = source_label

    input_type = collapse_spaces(clean_text(row["input_type"]))
    if input_type:
        field["input_type"] = input_type

    options = parse_option_lines(clean_text(row["options"]))
    if options:
        if canonical_key(input_type) == "predefined selection":
            field["dropdown_options"] = options
        elif len(options) == 1:
            field["value_hint"] = options[0]
        else:
            field["value_options"] = options

    normal_value = collapse_spaces(clean_text(row["normal_value"]))
    if normal_value and not is_placeholder(normal_value):
        field["normal_value"] = normal_value

    notes = collapse_spaces(clean_text(row["notes"]))
    if notes and not is_placeholder(notes):
        field["notes"] = notes

    return field


def find_section(sections: list[dict[str, object]], section_name: str) -> dict[str, object] | None:
    wanted = canonical_key(section_name)
    for section in sections:
        if canonical_key(str(section["name"])) == wanted:
            return section
    return None


def ensure_section(
    form: dict[str, object],
    section_name: str,
    sheet_row: int | None = None,
    source_label: str | None = None,
    merge_repeated: bool = False,
) -> dict[str, object]:
    sections = form["sections"]
    assert isinstance(sections, list)

    if merge_repeated:
        existing = find_section(sections, section_name)
        if existing is not None:
            return existing

    section = {
        "name": section_name,
        "fields": [],
        "notes": [],
    }
    if sheet_row is not None:
        section["sheet_row"] = sheet_row
    if source_label and source_label != section_name:
        section["source_label"] = source_label

    sections.append(section)
    return section


def move_fields_after(
    fields: list[dict[str, object]],
    rules: list[tuple[str, str]],
) -> None:
    for field_name, anchor_name in rules:
        field_key = canonical_key(field_name)
        anchor_key = canonical_key(anchor_name)

        field_index = next(
            (index for index, item in enumerate(fields) if canonical_key(str(item["name"])) == field_key),
            None,
        )
        anchor_index = next(
            (index for index, item in enumerate(fields) if canonical_key(str(item["name"])) == anchor_key),
            None,
        )

        if field_index is None or anchor_index is None or field_index == anchor_index + 1:
            continue

        field = fields.pop(field_index)
        if field_index < anchor_index:
            anchor_index -= 1
        fields.insert(anchor_index + 1, field)


def apply_field_order(form: dict[str, object], spec: dict[str, object]) -> None:
    rules = spec.get("move_fields_after")
    if not rules:
        return

    move_fields_after(form["fields"], list(rules))

    for section in form["sections"]:
        move_fields_after(section["fields"], list(rules))


def parse_form(worksheet, spec: dict[str, object]) -> dict[str, object]:
    rows, extra_notes = extract_rows(worksheet)
    synthetic_sections = spec.get("synthetic_sections", {})
    synthetic_section_map: dict[str, str] = {}
    for section_name, labels in synthetic_sections.items():
        for label in labels:
            synthetic_section_map[canonical_key(str(label))] = str(section_name)

    form = {
        "name": spec["form"],
        "sheet": worksheet.title,
        "notes": list(extra_notes),
        "fields": [],
        "sections": [],
    }
    current_section: dict[str, object] | None = None

    for row in rows:
        raw_label = clean_text(row["field"])
        label_key = canonical_key(raw_label)

        if label_key in COMMON_FIELDS:
            continue

        if not raw_label:
            inline_notes = combine_notes(str(row["normal_value"]), str(row["notes"]))
            if inline_notes:
                target = current_section["notes"] if current_section else form["notes"]
                assert isinstance(target, list)
                for note in inline_notes:
                    append_unique(target, note)
            continue

        if not clean_text(row["input_type"]) and not clean_text(row["options"]):
            section = ensure_section(
                form,
                section_name=display_label(raw_label),
                sheet_row=int(row["row"]),
                source_label=raw_label,
                merge_repeated=bool(spec.get("merge_repeated_sections")),
            )
            for note in combine_notes(str(row["normal_value"]), str(row["notes"])):
                append_unique(section["notes"], note)
            current_section = section
            continue

        field = build_field(row)
        synthetic_section_name = synthetic_section_map.get(label_key)
        if synthetic_section_name:
            section = ensure_section(
                form,
                section_name=synthetic_section_name,
                merge_repeated=True,
            )
            section["fields"].append(field)
            current_section = None
            continue

        if current_section is None:
            form["fields"].append(field)
        else:
            current_section["fields"].append(field)

    form["notes"] = [note for note in form["notes"] if note]
    for section in form["sections"]:
        section["notes"] = [note for note in section["notes"] if note]

    apply_field_order(form, spec)
    return form


def build_tree(workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, data_only=False)
    nodes: list[dict[str, object]] = []
    grouped_specs: OrderedDict[str, list[dict[str, object]]] = OrderedDict()

    for spec in FORM_SPECS:
        grouped_specs.setdefault(str(spec["group"]), []).append(spec)

    for group_name, specs in grouped_specs.items():
        first_spec = specs[0]
        group_kind = str(first_spec["group_kind"])

        if group_kind == "category":
            node = {
                "name": group_name,
                "kind": "category",
                "forms": [],
            }
            for spec in specs:
                worksheet = workbook[str(spec["sheet"])]
                node["forms"].append(parse_form(worksheet, spec))
        else:
            spec = specs[0]
            worksheet = workbook[str(spec["sheet"])]
            form = parse_form(worksheet, spec)
            node = {
                "name": group_name,
                "kind": "standalone_form",
                "sheet": form["sheet"],
                "notes": form["notes"],
                "fields": form["fields"],
                "sections": form["sections"],
            }

        nodes.append(node)

    return {
        "source_workbook": workbook_path.name,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "ignored_common_fields": [
            "Name",
            "Age",
            "Sex",
            "Date / Date-Time",
            "Requesting Physician",
            "Room",
            "Case Number",
            "Medical Technologist",
            "Pathologist",
        ],
        "tree": nodes,
    }


def looks_numeric_hint(value: str) -> bool:
    text = collapse_spaces(clean_text(value)).lower()
    return bool(text) and any(char.isdigit() for char in text)


def make_unique_key(value: str, used_keys: set[str]) -> str:
    base_key = slugify(value)
    key = base_key
    suffix = 2
    while key in used_keys:
        key = f"{base_key}_{suffix}"
        suffix += 1
    used_keys.add(key)
    return key


def get_field_option_values(field: dict[str, object]) -> list[str]:
    options: list[str] = []
    for key in ("dropdown_options", "value_options"):
        values = field.get(key)
        if not values:
            continue
        assert isinstance(values, list)
        options.extend(str(item) for item in values if str(item).strip())
    return options


def split_option_label_and_unit(value: str) -> tuple[str, str]:
    text = collapse_spaces(clean_text(value))
    if ":" not in text:
        return text, ""
    label, unit = text.split(":", 1)
    return collapse_spaces(label), collapse_spaces(unit)


def is_vital_signs_field_group(field: dict[str, object]) -> bool:
    if canonical_key(str(field.get("name", ""))) != "vital signs":
        return False

    option_names = [
        canonical_key(split_option_label_and_unit(option_value)[0])
        for option_value in get_field_option_values(field)
    ]
    return option_names == list(VITAL_SIGNS_CHILD_SPECS.keys())


def normalize_field_control(field: dict[str, object]) -> str:
    if get_field_option_values(field):
        return "select"
    return "input"


def infer_field_data_type(field: dict[str, object], control: str) -> str:
    if control == "select":
        return "enum"

    field_name = canonical_key(str(field.get("name", "")))
    if "date/time" in field_name or ("date" in field_name and "time" in field_name):
        return "datetime"
    if field_name.startswith("time ") or field_name == "time":
        return "time"
    if "date" in field_name:
        return "date"
    if field.get("value_hint") or looks_numeric_hint(str(field.get("normal_value", ""))):
        return "number"
    return "text"


def build_app_options(option_values: list[str], field_id: str) -> list[dict[str, object]]:
    options: list[dict[str, object]] = []
    used_keys: set[str] = set()
    for order, option_name in enumerate(option_values, start=1):
        key = make_unique_key(option_name, used_keys)
        options.append(
            {
                "id": f"{field_id}.{key}",
                "key": key,
                "name": option_name,
                "order": order,
            }
        )
    return options


def build_app_field(field: dict[str, object], parent_id: str, order: int, used_keys: set[str]) -> dict[str, object]:
    key = make_unique_key(str(field["name"]), used_keys)
    field_id = f"{parent_id}.{key}"
    control = normalize_field_control(field)
    data_type = infer_field_data_type(field, control)
    option_values = get_field_option_values(field)

    app_field: dict[str, object] = {
        "id": field_id,
        "key": key,
        "name": field["name"],
        "kind": "field",
        "order": order,
        "control": control,
        "data_type": data_type,
        "notes": [],
        "source": {
            "sheet_row": field.get("sheet_row"),
            "raw_input_type": field.get("input_type"),
        },
    }

    if field.get("source_label"):
        app_field["source"]["source_label"] = field["source_label"]

    value_hint = collapse_spaces(clean_text(field.get("value_hint")))
    if value_hint:
        app_field["unit_hint"] = value_hint

    normal_value = collapse_spaces(clean_text(field.get("normal_value")))
    if normal_value:
        app_field["normal_value"] = normal_value

    note = collapse_spaces(clean_text(field.get("notes")))
    if note:
        app_field["notes"].append(note)

    if option_values:
        app_field["options"] = build_app_options(option_values, field_id)

    if not app_field["notes"]:
        app_field.pop("notes")

    return app_field


def build_vital_signs_field_group(
    field: dict[str, object],
    parent_id: str,
    order: int,
    used_keys: set[str],
) -> dict[str, object]:
    key = make_unique_key(str(field["name"]), used_keys)
    group_id = f"{parent_id}.{key}"
    child_used_keys: set[str] = set()
    notes = []

    note = collapse_spaces(clean_text(field.get("notes")))
    if note:
        notes.append(note)

    child_fields: list[dict[str, object]] = []
    for child_order, option_value in enumerate(get_field_option_values(field), start=1):
        child_name, child_unit = split_option_label_and_unit(option_value)
        child_key = make_unique_key(child_name, child_used_keys)
        child_id = f"{group_id}.{child_key}"
        spec = VITAL_SIGNS_CHILD_SPECS[canonical_key(child_name)]

        child_field: dict[str, object] = {
            "id": child_id,
            "key": child_key,
            "name": child_name,
            "kind": "field",
            "order": child_order,
            "control": "input",
            "data_type": spec["data_type"],
            "source": {
                "sheet_row": field.get("sheet_row"),
                "raw_input_type": field.get("input_type"),
                "normalized_from_option": option_value,
            },
        }

        unit_hint = child_unit or str(spec.get("unit_hint", "")).strip()
        if unit_hint:
            child_field["unit_hint"] = unit_hint

        child_fields.append(child_field)

    group: dict[str, object] = {
        "id": group_id,
        "key": key,
        "name": field["name"],
        "kind": "field_group",
        "order": order,
        "source": {
            "sheet_row": field.get("sheet_row"),
            "raw_input_type": field.get("input_type"),
            "normalized_from": "option_list_to_field_group",
        },
        "fields": child_fields,
    }

    if notes:
        group["notes"] = notes

    return group


def build_app_section_child(
    field: dict[str, object],
    parent_id: str,
    order: int,
    used_keys: set[str],
) -> dict[str, object]:
    if is_vital_signs_field_group(field):
        return build_vital_signs_field_group(field, parent_id, order, used_keys)
    return build_app_field(field, parent_id, order, used_keys)


def build_app_section(section: dict[str, object], form_id: str, order: int) -> dict[str, object]:
    key = slugify(str(section["name"]))
    section_id = f"{form_id}.{key}"
    used_field_keys: set[str] = set()
    app_section: dict[str, object] = {
        "id": section_id,
        "key": key,
        "name": section["name"],
        "order": order,
        "notes": [note for note in section.get("notes", []) if str(note).strip()],
        "source": {
            "sheet_row": section.get("sheet_row"),
        },
        "fields": [
            build_app_section_child(field, section_id, field_order, used_field_keys)
            for field_order, field in enumerate(section.get("fields", []), start=1)
        ],
    }

    if section.get("source_label"):
        app_section["source"]["source_label"] = section["source_label"]
    if not app_section["notes"]:
        app_section.pop("notes")

    return app_section


def build_app_form(form: dict[str, object], group_id: str, order: int) -> dict[str, object]:
    form_key = slugify(str(form["name"]))
    form_id = f"{group_id}.{form_key}"
    used_field_keys: set[str] = set()

    return {
        "id": form_id,
        "key": form_key,
        "name": form["name"],
        "order": order,
        "common_field_set_id": COMMON_FIELD_SET["id"],
        "notes": [note for note in form.get("notes", []) if str(note).strip()],
        "source": {
            "sheet": form.get("sheet"),
        },
        "fields": [
            build_app_field(field, form_id, field_order, used_field_keys)
            for field_order, field in enumerate(form.get("fields", []), start=1)
        ],
        "sections": [
            build_app_section(section, form_id, section_order)
            for section_order, section in enumerate(form.get("sections", []), start=1)
        ],
    }


def build_app_group(node: dict[str, object], order: int) -> dict[str, object]:
    group_id = slugify(str(node["name"]))
    if node["kind"] == "category":
        forms = [
            build_app_form(form, group_id, form_order)
            for form_order, form in enumerate(node.get("forms", []), start=1)
        ]
    else:
        standalone_form = {
            "name": node["name"],
            "sheet": node.get("sheet"),
            "notes": node.get("notes", []),
            "fields": node.get("fields", []),
            "sections": node.get("sections", []),
        }
        forms = [build_app_form(standalone_form, group_id, 1)]

    return {
        "id": group_id,
        "name": node["name"],
        "kind": node["kind"],
        "order": order,
        "forms": forms,
    }


def build_app_schema(document: dict[str, object]) -> dict[str, object]:
    return {
        "schema_name": "NAIC Medtech App Schema",
        "schema_version": APP_SCHEMA_VERSION,
        "source_workbook": document["source_workbook"],
        "generated_at_utc": document["generated_at_utc"],
        "normalization_rules": [
            "Shared patient/request metadata was moved into a reusable common field set.",
            "Medical Technologist and Pathologist are modeled as configurable print signatories, not ordinary result fields.",
            'Any field with explicit options was normalized to control="select", even if the raw spreadsheet input type said manual entry.',
            "Known pseudo-option lists that actually represent child manual-entry fields were normalized into field groups.",
            'Manual-entry fields with strong numeric hints such as units or numeric normal ranges were normalized to data_type="number".',
            'Manual-entry fields without strong type signals default to data_type="text".',
            "Standalone departments are represented as groups that contain exactly one form for a consistent app navigation model.",
        ],
        "common_field_sets": [COMMON_FIELD_SET],
        "default_signatories": DEFAULT_SIGNATORIES,
        "groups": [
            build_app_group(node, group_order)
            for group_order, node in enumerate(document.get("tree", []), start=1)
        ],
    }


def render_field(field: dict[str, object], indent: str) -> list[str]:
    lines = [f"{indent}{field['name']}"]

    scalar_keys = [
        ("input_type", "input_type"),
        ("value_hint", "value_hint"),
        ("normal_value", "normal_value"),
        ("notes", "notes"),
    ]
    list_keys = [
        ("dropdown_options", "dropdown_options"),
        ("value_options", "value_options"),
    ]

    for key, label in scalar_keys:
        value = field.get(key)
        if value:
            lines.append(f"{indent}  {label}: {value}")

    for key, label in list_keys:
        values = field.get(key)
        if values:
            lines.append(f"{indent}  {label}:")
            for item in values:
                lines.append(f"{indent}    - {item}")

    return lines


def render_section(section: dict[str, object], indent: str) -> list[str]:
    lines = [f"{indent}{section['name']}"]

    for note in section.get("notes", []):
        lines.append(f"{indent}  note: {note}")

    for field in section.get("fields", []):
        lines.extend(render_field(field, indent + "  "))

    return lines


def render_form(form: dict[str, object], indent: str) -> list[str]:
    lines = [f"{indent}{form['name']}"]

    lines.extend(render_form_body(form, indent + "  "))
    return lines


def render_form_body(form: dict[str, object], indent: str) -> list[str]:
    lines: list[str] = []

    for note in form.get("notes", []):
        lines.append(f"{indent}note: {note}")

    for field in form.get("fields", []):
        lines.extend(render_field(field, indent))

    for section in form.get("sections", []):
        lines.extend(render_section(section, indent))

    return lines


def build_markdown(document: dict[str, object]) -> str:
    lines = [
        "# NAIC Medtech Structured Tree",
        "",
        "Client-approved grouping was used for categories/forms.",
        "Common/shared fields were intentionally excluded from the tree below.",
        "",
        "Ignored common fields:",
    ]

    for field_name in document["ignored_common_fields"]:
        lines.append(f"- {field_name}")

    lines.append("")
    lines.append("```text")

    for node in document["tree"]:
        lines.append(str(node["name"]))
        if node["kind"] == "category":
            for form in node["forms"]:
                lines.extend(render_form(form, "  "))
        else:
            standalone_form = {
                "notes": node.get("notes", []),
                "fields": node.get("fields", []),
                "sections": node.get("sections", []),
            }
            lines.extend(render_form_body(standalone_form, "  "))
        lines.append("")

    lines.append("```")
    lines.append("")
    return "\n".join(lines)


def build_html(document: dict[str, object]) -> str:
    tree_json = json.dumps(document["tree"], ensure_ascii=False)
    ignored_json = json.dumps(document["ignored_common_fields"], ensure_ascii=False)

    template = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>NAIC Medtech Structured Tree</title>
  <style>
    :root {
      --bg: #f5f1e8;
      --panel: rgba(255, 250, 240, 0.92);
      --panel-strong: #fffdf8;
      --line: #d7cbb6;
      --text: #2e2419;
      --muted: #6a5b4a;
      --accent: #9c4f2d;
      --accent-soft: #f0d3bf;
      --chip: #efe4d2;
      --shadow: 0 18px 40px rgba(84, 61, 34, 0.08);
    }

    * {
      box-sizing: border-box;
    }

    body {
      margin: 0;
      font-family: "Segoe UI", "Trebuchet MS", sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(156, 79, 45, 0.12), transparent 28%),
        linear-gradient(180deg, #fbf7f0 0%, var(--bg) 100%);
    }

    .shell {
      width: min(1200px, calc(100vw - 32px));
      margin: 24px auto 40px;
    }

    .hero {
      background: linear-gradient(135deg, rgba(255,255,255,0.92), rgba(248, 239, 226, 0.94));
      border: 1px solid rgba(156, 79, 45, 0.14);
      border-radius: 24px;
      padding: 24px;
      box-shadow: var(--shadow);
    }

    h1 {
      margin: 0 0 8px;
      font-size: clamp(28px, 3vw, 42px);
      line-height: 1.05;
      letter-spacing: -0.03em;
    }

    .subtitle {
      margin: 0;
      color: var(--muted);
      max-width: 78ch;
    }

    .toolbar {
      margin-top: 18px;
      display: grid;
      grid-template-columns: minmax(280px, 1fr) auto auto;
      gap: 12px;
      align-items: center;
    }

    .toolbar input {
      width: 100%;
      border: 1px solid rgba(156, 79, 45, 0.2);
      background: rgba(255, 255, 255, 0.86);
      border-radius: 14px;
      padding: 13px 15px;
      font-size: 15px;
      color: var(--text);
      outline: none;
    }

    .toolbar input:focus {
      border-color: rgba(156, 79, 45, 0.45);
      box-shadow: 0 0 0 4px rgba(156, 79, 45, 0.12);
    }

    button {
      border: 0;
      background: var(--text);
      color: #fff;
      border-radius: 999px;
      padding: 11px 16px;
      font-weight: 600;
      cursor: pointer;
    }

    button.secondary {
      background: #7f6a50;
    }

    .meta {
      margin-top: 16px;
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }

    .pill {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      padding: 7px 12px;
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.7);
      border: 1px solid rgba(156, 79, 45, 0.14);
      color: var(--muted);
      font-size: 13px;
    }

    .legend {
      margin-top: 10px;
      color: var(--muted);
      font-size: 13px;
    }

    .ignored {
      margin-top: 10px;
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }

    .ignored .chip {
      background: var(--chip);
      color: var(--text);
      border-radius: 999px;
      padding: 6px 10px;
      font-size: 12px;
    }

    .grid {
      margin-top: 22px;
      display: grid;
      gap: 16px;
    }

    .group-card {
      background: var(--panel);
      border: 1px solid rgba(156, 79, 45, 0.12);
      border-radius: 22px;
      padding: 18px 18px 10px;
      box-shadow: var(--shadow);
    }

    .group-head {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      margin-bottom: 10px;
    }

    .group-head h2 {
      margin: 0;
      font-size: 22px;
      letter-spacing: -0.02em;
    }

    .count {
      color: var(--muted);
      font-size: 13px;
      background: rgba(156, 79, 45, 0.08);
      padding: 6px 10px;
      border-radius: 999px;
    }

    .tree-stack {
      display: grid;
      gap: 12px;
      padding-bottom: 8px;
    }

    details.tree-node {
      border-left: 2px solid var(--line);
      margin-left: 8px;
      padding-left: 14px;
    }

    details.tree-node > summary {
      list-style: none;
      cursor: pointer;
      padding: 8px 10px;
      border-radius: 12px;
      background: rgba(255, 255, 255, 0.56);
      border: 1px solid rgba(156, 79, 45, 0.08);
      user-select: none;
    }

    details.tree-node > summary::-webkit-details-marker {
      display: none;
    }

    details.tree-node > summary::before {
      content: "+";
      display: inline-flex;
      justify-content: center;
      align-items: center;
      width: 20px;
      height: 20px;
      border-radius: 999px;
      margin-right: 10px;
      background: var(--accent-soft);
      color: var(--accent);
      font-weight: 800;
    }

    details.tree-node[open] > summary::before {
      content: "-";
    }

    .node-body {
      margin: 10px 0 4px 2px;
      display: grid;
      gap: 10px;
    }

    .note {
      padding: 9px 12px;
      border-radius: 12px;
      background: rgba(156, 79, 45, 0.08);
      color: var(--text);
      font-size: 13px;
    }

    .field {
      background: var(--panel-strong);
      border: 1px solid rgba(156, 79, 45, 0.1);
      border-radius: 16px;
      padding: 12px 14px;
      box-shadow: 0 8px 18px rgba(84, 61, 34, 0.04);
    }

    .field-title {
      font-weight: 700;
      letter-spacing: -0.01em;
    }

    .field-meta {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 8px;
    }

    .field-meta .chip {
      padding: 5px 9px;
      background: var(--chip);
      border-radius: 10px;
      font-size: 12px;
      color: var(--text);
    }

    .option-list {
      margin-top: 10px;
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }

    .option {
      padding: 6px 10px;
      border-radius: 999px;
      background: rgba(156, 79, 45, 0.08);
      border: 1px solid rgba(156, 79, 45, 0.12);
      font-size: 12px;
      color: var(--text);
    }

    .empty-state {
      padding: 24px;
      border-radius: 18px;
      text-align: center;
      color: var(--muted);
      background: rgba(255, 255, 255, 0.6);
      border: 1px dashed rgba(156, 79, 45, 0.22);
    }

    .hidden {
      display: none !important;
    }

    .match > summary,
    .field.match {
      outline: 2px solid rgba(156, 79, 45, 0.36);
      outline-offset: 2px;
    }

    @media (max-width: 760px) {
      .toolbar {
        grid-template-columns: 1fr;
      }

      button {
        width: 100%;
      }
    }
  </style>
</head>
<body>
  <main class="shell">
    <section class="hero">
      <h1>NAIC Medtech Structured Tree</h1>
      <p class="subtitle">
        Interactive tree view generated from the Excel workbook using the client-approved grouping.
        Shared patient/staff metadata fields are excluded so you can focus on the unique structure of each form.
      </p>

      <div class="toolbar">
        <input id="searchInput" type="search" placeholder="Search category, form, section, field, option, note...">
        <button id="expandAllBtn" type="button">Expand All</button>
        <button id="collapseAllBtn" class="secondary" type="button">Collapse All</button>
      </div>

      <div class="meta">
        <span class="pill" id="groupCount"></span>
        <span class="pill" id="formCount"></span>
      </div>

      <div class="legend">
        Tip: use search to isolate one lab form, test name, dropdown option, or note. Matching branches auto-expand.
      </div>

      <div class="ignored" id="ignoredCommonFields"></div>
    </section>

    <section class="grid" id="treeRoot"></section>
  </main>

  <script>
    const treeData = __TREE_DATA__;
    const ignoredCommonFields = __IGNORED_FIELDS__;

    const treeRoot = document.getElementById("treeRoot");
    const ignoredRoot = document.getElementById("ignoredCommonFields");
    const searchInput = document.getElementById("searchInput");
    const expandAllBtn = document.getElementById("expandAllBtn");
    const collapseAllBtn = document.getElementById("collapseAllBtn");
    const groupCount = document.getElementById("groupCount");
    const formCount = document.getElementById("formCount");

    function appendChild(parent, child) {
      if (child) parent.appendChild(child);
      return child;
    }

    function createElement(tag, className, text) {
      const el = document.createElement(tag);
      if (className) el.className = className;
      if (text !== undefined) el.textContent = text;
      return el;
    }

    function createChip(text) {
      return createElement("span", "chip", text);
    }

    function aggregateFieldText(field) {
      const parts = [
        field.name,
        field.input_type,
        field.value_hint,
        field.normal_value,
        field.notes,
        ...(field.dropdown_options || []),
        ...(field.value_options || [])
      ];
      return parts.filter(Boolean).join(" ").toLowerCase();
    }

    function aggregateSectionText(section) {
      const parts = [section.name, ...(section.notes || [])];
      for (const field of section.fields || []) {
        parts.push(aggregateFieldText(field));
      }
      return parts.filter(Boolean).join(" ").toLowerCase();
    }

    function aggregateFormText(form) {
      const parts = [form.name, ...(form.notes || [])];
      for (const field of form.fields || []) {
        parts.push(aggregateFieldText(field));
      }
      for (const section of form.sections || []) {
        parts.push(aggregateSectionText(section));
      }
      return parts.filter(Boolean).join(" ").toLowerCase();
    }

    function createOptionList(values) {
      if (!values || !values.length) return null;
      const wrap = createElement("div", "option-list");
      for (const value of values) {
        appendChild(wrap, createElement("span", "option", value));
      }
      return wrap;
    }

    function createField(field) {
      const card = createElement("article", "field searchable");
      card.dataset.search = aggregateFieldText(field);

      appendChild(card, createElement("div", "field-title", field.name));

      const meta = createElement("div", "field-meta");
      let hasMeta = false;

      if (field.input_type) {
        appendChild(meta, createChip("input: " + field.input_type));
        hasMeta = true;
      }
      if (field.value_hint) {
        appendChild(meta, createChip("hint: " + field.value_hint));
        hasMeta = true;
      }
      if (field.normal_value) {
        appendChild(meta, createChip("normal: " + field.normal_value));
        hasMeta = true;
      }
      if (field.notes) {
        appendChild(meta, createChip("note: " + field.notes));
        hasMeta = true;
      }

      if (hasMeta) appendChild(card, meta);
      appendChild(card, createOptionList(field.dropdown_options));
      appendChild(card, createOptionList(field.value_options));

      return card;
    }

    function createBranch(name, searchText, notes, fields, sections, openByDefault) {
      const details = createElement("details", "tree-node searchable");
      details.open = Boolean(openByDefault);
      details.dataset.search = searchText;

      appendChild(details, createElement("summary", "", name));

      const body = createElement("div", "node-body");
      for (const note of notes || []) {
        appendChild(body, createElement("div", "note", note));
      }
      for (const field of fields || []) {
        appendChild(body, createField(field));
      }
      for (const section of sections || []) {
        appendChild(body, createSection(section));
      }
      appendChild(details, body);

      return details;
    }

    function createSection(section) {
      return createBranch(
        section.name,
        aggregateSectionText(section),
        section.notes || [],
        section.fields || [],
        [],
        false
      );
    }

    function createForm(form) {
      return createBranch(
        form.name,
        aggregateFormText(form),
        form.notes || [],
        form.fields || [],
        form.sections || [],
        true
      );
    }

    function createStandaloneBody(form) {
      const body = createElement("div", "node-body");
      for (const note of form.notes || []) {
        appendChild(body, createElement("div", "note", note));
      }
      for (const field of form.fields || []) {
        appendChild(body, createField(field));
      }
      for (const section of form.sections || []) {
        appendChild(body, createSection(section));
      }
      return body;
    }

    function renderTree() {
      let totalForms = 0;

      for (const item of ignoredCommonFields) {
        appendChild(ignoredRoot, createElement("span", "chip", item));
      }

      for (const node of treeData) {
        const card = createElement("section", "group-card searchable");
        const head = createElement("div", "group-head");
        const title = createElement("h2", "", node.name);
        const badgeText = node.kind === "category"
          ? (node.forms.length + " form" + (node.forms.length > 1 ? "s" : ""))
          : "standalone form";
        const badge = createElement("span", "count", badgeText);

        appendChild(head, title);
        appendChild(head, badge);
        appendChild(card, head);

        const stack = createElement("div", "tree-stack");
        if (node.kind === "category") {
          totalForms += node.forms.length;
          node.forms.forEach(form => appendChild(stack, createForm(form)));
          card.dataset.search = (node.name + " " + node.forms.map(aggregateFormText).join(" ")).toLowerCase();
        } else {
          totalForms += 1;
          const standaloneForm = {
            name: node.name,
            notes: node.notes || [],
            fields: node.fields || [],
            sections: node.sections || []
          };
          appendChild(stack, createStandaloneBody(standaloneForm));
          card.dataset.search = (node.name + " " + aggregateFormText(standaloneForm)).toLowerCase();
        }

        appendChild(card, stack);
        appendChild(treeRoot, card);
      }

      groupCount.textContent = treeData.length + " top-level groups";
      formCount.textContent = totalForms + " total forms";
    }

    function setAllDetails(open) {
      document.querySelectorAll("details.tree-node").forEach(node => {
        node.open = open;
      });
    }

    function filterTree(query) {
      const normalized = query.trim().toLowerCase();
      const nodes = Array.from(document.querySelectorAll(".searchable"));
      const cards = Array.from(document.querySelectorAll(".group-card"));
      let anyVisible = false;

      nodes.forEach(node => {
        node.classList.remove("hidden", "match");
      });

      if (!normalized) {
        cards.forEach(card => card.classList.remove("hidden"));
        const emptyState = document.getElementById("emptyState");
        if (emptyState) emptyState.remove();
        return;
      }

      nodes.forEach(node => {
        const haystack = node.dataset.search || "";
        const matches = haystack.includes(normalized);
        if (!matches) node.classList.add("hidden");
        if (matches) node.classList.add("match");
      });

      document.querySelectorAll("details.tree-node.match").forEach(node => {
        node.open = true;
      });

      cards.forEach(card => {
        const visibleDescendant = card.querySelector(".searchable:not(.hidden)");
        if (!visibleDescendant) {
          card.classList.add("hidden");
        } else {
          card.classList.remove("hidden");
          anyVisible = true;
        }
      });

      let emptyState = document.getElementById("emptyState");
      if (!anyVisible) {
        if (!emptyState) {
          emptyState = createElement("div", "empty-state");
          emptyState.id = "emptyState";
          emptyState.textContent = "No matching nodes found. Try a broader search term.";
          treeRoot.appendChild(emptyState);
        }
      } else if (emptyState) {
        emptyState.remove();
      }
    }

    renderTree();

    searchInput.addEventListener("input", event => {
      filterTree(event.target.value);
    });

    expandAllBtn.addEventListener("click", () => setAllDetails(true));
    collapseAllBtn.addEventListener("click", () => setAllDetails(false));
  </script>
</body>
</html>
"""

    return (
        template.replace("__TREE_DATA__", tree_json).replace("__IGNORED_FIELDS__", ignored_json)
    )


def build_diagram_html(app_schema: dict[str, object]) -> str:
    schema_json = json.dumps(app_schema, ensure_ascii=False)

    template = """<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>NAIC Medtech Single Tree</title><style>
body{margin:0;font-family:"Segoe UI","Trebuchet MS",sans-serif;color:#2f2419;background:linear-gradient(180deg,#faf7f1,#f4efe7)}
.shell{width:min(1600px,calc(100vw - 24px));margin:18px auto 32px}
.hero,.panel{background:rgba(255,252,246,.96);border:1px solid rgba(47,36,25,.10);border-radius:24px;box-shadow:0 18px 40px rgba(67,47,26,.10)}
.hero{padding:22px}.panel{margin-top:18px;padding:16px}
h1{margin:0 0 8px;font-size:clamp(30px,4vw,46px);line-height:1.02;letter-spacing:-.03em}
.subtitle,.helper,.foot{color:#6f6253}.helper,.foot{font-size:13px}.helper{margin-top:12px}.foot{margin-top:14px}
.toolbar{margin-top:18px;display:grid;grid-template-columns:minmax(280px,1fr) auto auto;gap:10px;align-items:center}
.toolbar input{width:100%;border:1px solid rgba(47,36,25,.16);border-radius:14px;padding:12px 14px;background:#fff;font-size:15px;color:#2f2419;outline:none}
.toolbar input:focus{border-color:rgba(47,36,25,.34);box-shadow:0 0 0 4px rgba(47,36,25,.10)}
.toolbar button{border:0;border-radius:999px;padding:11px 15px;background:#2f2419;color:#fff;font-weight:700;font-size:14px;cursor:pointer}.toolbar .secondary{background:#77634c}
.scroll{overflow:auto;min-height:78vh;padding:10px 12px 22px;border-radius:18px;background:linear-gradient(180deg,rgba(255,255,255,.45),rgba(255,255,255,.25)),repeating-linear-gradient(90deg,rgba(47,36,25,.02) 0,rgba(47,36,25,.02) 1px,transparent 1px,transparent 22px),repeating-linear-gradient(180deg,rgba(47,36,25,.02) 0,rgba(47,36,25,.02) 1px,transparent 1px,transparent 22px)}
.tree,.tree ul{margin:0;padding:0;list-style:none;display:flex;justify-content:center;position:relative;width:max-content;min-width:100%}
.tree ul{padding-top:26px}.item{position:relative;padding:26px 5px 0;text-align:center}.item:before,.item:after{content:"";position:absolute;top:0;width:50%;height:26px;border-top:2px solid #30261d}.item:before{right:50%}.item:after{left:50%;border-left:2px solid #30261d}.item:only-child:before,.item:only-child:after{display:none}.item:only-child{padding-top:0}.item:first-child:before,.item:last-child:after{border:0}.item:last-child:before{border-right:2px solid #30261d;border-radius:0 8px 0 0}.item:first-child:after{border-radius:8px 0 0 0}.item>ul:before{content:"";position:absolute;top:0;left:50%;height:26px;border-left:2px solid #30261d}.item.collapsed>ul,.item.hidden{display:none}
.btn{border:1px solid rgba(47,36,25,.18);border-radius:14px;box-shadow:0 10px 22px rgba(67,47,26,.08);min-width:118px;max-width:154px;min-height:54px;padding:9px 10px;display:inline-flex;gap:8px;align-items:center;justify-content:space-between;background:#fff;color:#2f2419;text-align:center;font:600 12px/1.2 "Segoe UI","Trebuchet MS",sans-serif;cursor:pointer}
.kind-root>.btn{min-width:186px;background:#e7f5c8}.kind-category>.btn{background:#ffdcbf}.kind-form>.btn{background:#cfe0ff}.kind-section>.btn{background:#f7e7ad}.kind-field_group>.btn{background:#f1e9dd}.kind-option>.btn{background:#e7edf8;font-weight:600}
.content{display:flex;flex-direction:column;align-items:center;justify-content:center;gap:3px;flex:1 1 auto}.meta{font-size:9px;line-height:1;letter-spacing:.12em;text-transform:uppercase;color:rgba(47,36,25,.55);font-weight:700}.label{font-size:12px;flex:1 1 auto}.sub,.normal{font-size:10px;line-height:1.18;color:rgba(47,36,25,.68);font-weight:600}.normal{font-weight:500;color:rgba(47,36,25,.62)}.kind-root>.btn .label{font-size:14px}.kind-field>.btn{min-height:72px;max-width:178px}.toggle{width:22px;height:22px;border-radius:999px;display:inline-flex;align-items:center;justify-content:center;background:rgba(47,36,25,.10);font-size:15px;line-height:1;flex:0 0 auto}.no-children .toggle{display:none}.no-children>.btn{cursor:default}.matched>.btn{outline:3px solid rgba(155,77,44,.35);outline-offset:2px}
@media (max-width:1100px){.toolbar{grid-template-columns:1fr}.toolbar button{width:100%}}
</style></head><body><main class="shell"><section class="hero"><h1>Single Tree Diagram</h1><p class="subtitle">One literal tree only. Shared/common fields are excluded, and subtle field hints like data type and normal value are shown only where they matter.</p><div class="toolbar"><input id="q" type="search" placeholder="Search category, form, section, field, or option"><button id="expand" type="button">Expand All</button><button id="collapse" class="secondary" type="button">Collapse All</button></div><div class="helper">Click any box with a plus sign to reveal its children. Grouped manual-entry fields like Vital Signs now appear as their own branch.</div></section><section class="panel"><div class="scroll" id="scroll"><div id="mount"></div></div><div class="foot">Root -> category / standalone form -> form -> section -> field group -> field -> option</div></section></main><script>
const schemaData=__SCHEMA_DATA__,mount=document.getElementById("mount"),scroll=document.getElementById("scroll"),q=document.getElementById("q");
const el=(t,c,x)=>{const n=document.createElement(t);if(c)n.className=c;if(x!==undefined)n.textContent=x;return n},add=(p,c)=>(c&&p.appendChild(c),c);
const optionNode=o=>({label:o.name,kind:"option",children:[]});
const fieldNode=f=>{if(f.kind==="field_group"){return{label:f.name,kind:"field_group",children:(f.fields||[]).map(fieldNode)}};const options=f.options||[],base=f.control==="select"?"choice":f.data_type||"value",sub=f.unit_hint?base+" • "+f.unit_hint:base;return{label:f.name,kind:"field",sub,normal:f.normal_value||"",children:options.map(optionNode)}};
const sectionNode=s=>({label:s.name,kind:"section",children:(s.fields||[]).map(fieldNode)});
const formNode=f=>({label:f.name,kind:"form",children:[...(f.fields||[]).map(fieldNode),...(f.sections||[]).map(sectionNode)]});
const topNode=g=>g.kind==="category"?{label:g.name,kind:"category",children:(g.forms||[]).map(formNode)}:formNode((g.forms||[])[0]||{name:g.name,fields:[],sections:[]});
const enrich=n=>{const kids=(n.children||[]).map(enrich),parts=[n.label,n.sub||"",n.normal||"",...kids.map(k=>k.search)];return {...n,children:kids,search:parts.join(" ").toLowerCase()}};
const treeData=enrich({label:"NAIC Medtech",kind:"root",children:(schemaData.groups||[]).map(topNode)});
function sync(li){const btn=li.querySelector(":scope>.btn"),tog=li.querySelector(":scope>.btn>.toggle"),has=!li.classList.contains("no-children");if(!btn||!tog)return;btn.setAttribute("aria-expanded",!has||!li.classList.contains("collapsed")?"true":"false");if(has)tog.textContent=li.classList.contains("collapsed")?"+":"-";}
function item(node,d){const li=el("li","item kind-"+node.kind);li.dataset.search=node.search;const has=node.children&&node.children.length;if(!has)li.classList.add("no-children");else if(d>0)li.classList.add("collapsed");const btn=el("button","btn");btn.type="button";const content=el("span","content");add(content,el("span","meta",node.kind));add(content,el("span","label",node.label));if(node.sub)add(content,el("span","sub",node.sub));if(node.normal)add(content,el("span","normal","normal: "+node.normal));add(btn,content);add(btn,el("span","toggle",has?"+":""));add(li,btn);if(has){const ul=el("ul");node.children.forEach(c=>add(ul,item(c,d+1)));add(li,ul);btn.addEventListener("click",()=>{li.classList.toggle("collapsed");sync(li);});}else btn.addEventListener("click",e=>e.preventDefault());sync(li);return li;}
function render(){mount.innerHTML="";const root=el("ul","tree");add(root,item(treeData,0));add(mount,root);const rb=mount.querySelector(".kind-root>.btn");if(rb){const t=rb.offsetLeft+rb.offsetWidth/2-scroll.clientWidth/2;scroll.scrollLeft=Math.max(0,t);}}
function setAll(c){mount.querySelectorAll(".item").forEach(i=>{if(i.classList.contains("no-children")||i.classList.contains("kind-root"))return;i.classList.toggle("collapsed",c);sync(i);});}
function clearState(){mount.querySelectorAll(".item").forEach(i=>i.classList.remove("hidden","matched"));}
function filterItem(i,s){const kids=(i.querySelector(":scope>ul")?Array.from(i.querySelector(":scope>ul").children):[]);let child=false;kids.forEach(k=>{if(filterItem(k,s))child=true;});const self=(i.dataset.search||"").includes(s),show=self||child;i.classList.toggle("hidden",!show);i.classList.toggle("matched",self);if(child&&!i.classList.contains("no-children")){i.classList.remove("collapsed");sync(i);}return show;}
function filterTree(v){const s=v.trim().toLowerCase();clearState();if(!s)return;const root=mount.querySelector(".tree>.item");if(root)filterItem(root,s);}
render();q.addEventListener("input",e=>filterTree(e.target.value));document.getElementById("expand").addEventListener("click",()=>{setAll(false);filterTree(q.value)});document.getElementById("collapse").addEventListener("click",()=>{setAll(true);filterTree(q.value)});
</script></body></html>"""

    return template.replace("__SCHEMA_DATA__", schema_json)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    document = build_tree(WORKBOOK_PATH)
    app_schema = build_app_schema(document)

    JSON_OUTPUT_PATH.write_text(
        json.dumps(document, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    APP_SCHEMA_OUTPUT_PATH.write_text(
        json.dumps(app_schema, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    MARKDOWN_OUTPUT_PATH.write_text(build_markdown(document), encoding="utf-8")
    HTML_OUTPUT_PATH.write_text(build_html(document), encoding="utf-8")
    DIAGRAM_HTML_OUTPUT_PATH.write_text(build_diagram_html(app_schema), encoding="utf-8")

    print(f"Wrote {JSON_OUTPUT_PATH}")
    print(f"Wrote {APP_SCHEMA_OUTPUT_PATH}")
    print(f"Wrote {MARKDOWN_OUTPUT_PATH}")
    print(f"Wrote {HTML_OUTPUT_PATH}")
    print(f"Wrote {DIAGRAM_HTML_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
